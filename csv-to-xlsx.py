#!/usr/bin/env python3
"""
csv-to-xlsx.py
Converts influencer master sheet FINAL.csv → influencer master sheet FINAL.xlsx
Using Python's csv module (correct RFC 4180 parsing) + openpyxl for formatting.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_CSV  = "influencer master sheet FINAL.csv"
OUTPUT_XLSX = "influencer master sheet FINAL.xlsx"

# ── Read CSV ──────────────────────────────────────────────────────────────────
with open(INPUT_CSV, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    all_rows = list(reader)

# Row 0: headers, Rows 1+: data (title row was removed earlier)
headers  = all_rows[0]   # ['Name', 'Social Platform URL', 'Handle', ...]
data_rows = all_rows[1:] # data

print(f"Headers: {headers}")
print(f"Data rows: {len(data_rows)}")

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Influencer Master Sheet"

# Column widths (in Excel character units)
COL_WIDTHS = [28, 52, 26, 34, 70, 55, 16, 16]

# ── Style constants ───────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F2328")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(vertical="center", horizontal="left", wrap_text=False)

LIGHT_FILL    = PatternFill("solid", fgColor="F7F8FA")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT     = Font(name="Calibri", size=11)
DATA_ALIGN    = Alignment(vertical="center", wrap_text=False)
URL_FONT      = Font(name="Calibri", size=11, color="3B82D4", underline="single")

BOTTOM_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))

# ── Header row ────────────────────────────────────────────────────────────────
ws.append(headers)
header_row = ws[1]
for cell in header_row:
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border    = BOTTOM_BORDER
ws.row_dimensions[1].height = 22

# Freeze header row
ws.freeze_panes = "A2"

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, row_data in enumerate(data_rows):
    if not any(cell.strip() for cell in row_data):
        continue  # skip blank rows

    # Pad row to 8 columns
    while len(row_data) < 8:
        row_data.append("")

    ws.append(row_data[:8])
    excel_row_num = ws.max_row
    excel_row = ws[excel_row_num]
    fill = WHITE_FILL if row_idx % 2 == 0 else LIGHT_FILL

    for col_idx, cell in enumerate(excel_row):
        cell.fill      = fill
        cell.alignment = DATA_ALIGN
        cell.font      = DATA_FONT

        # Make URL column a clickable hyperlink
        val = str(cell.value or "").strip()
        if col_idx == 1 and val.startswith("http"):
            # Take only first URL if multi-line
            first_url = val.split("\n")[0].strip()
            cell.hyperlink = first_url
            cell.value     = first_url
            cell.font      = URL_FONT

    ws.row_dimensions[excel_row_num].height = 18

# ── Column widths ─────────────────────────────────────────────────────────────
for col_idx, width in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Auto-filter (tab 1) ───────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# ── Tab 2: rows with empty URL ────────────────────────────────────────────────
ws2 = wb.create_sheet(title="Missing URL")

# Header row on tab 2
ws2.append(headers)
hdr2 = ws2[1]
for cell in hdr2:
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border    = BOTTOM_BORDER
ws2.row_dimensions[1].height = 22
ws2.freeze_panes = "A2"

# Rows where URL cell is blank / nan / not a real URL
no_url_rows = []
for r in data_rows:
    if not any(cell.strip() for cell in r):
        continue  # skip fully blank rows
    while len(r) < 8:
        r.append("")
    url = r[1].strip()
    if not url or url.lower() in ("nan", "-") or not url.startswith("http"):
        no_url_rows.append(r)

for row_idx, row_data in enumerate(no_url_rows):
    ws2.append(row_data[:8])
    excel_row_num = ws2.max_row
    excel_row = ws2[excel_row_num]
    fill = WHITE_FILL if row_idx % 2 == 0 else LIGHT_FILL
    for col_idx, cell in enumerate(excel_row):
        cell.fill      = fill
        cell.alignment = DATA_ALIGN
        cell.font      = DATA_FONT
    ws2.row_dimensions[excel_row_num].height = 18

# Column widths on tab 2
for col_idx, width in enumerate(COL_WIDTHS, 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

# Auto-filter on tab 2
ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws2.max_row}"

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
final_row  = ws.max_row  - 1
final_row2 = ws2.max_row - 1
print(f"Written: {OUTPUT_XLSX}  ({final_row} rows in tab 1, {final_row2} rows in tab 2 — Missing URL)")
